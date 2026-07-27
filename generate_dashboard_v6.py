import openpyxl
import re
from html import escape

f_q1 = 'C:/Users/ykqin3/Documents/Loomy Workspace/.loomy-attachments/ses_071ded120ffeOMnlYKgOSIt1Xp/1785086444217-0-_-2026_Q1_.xlsx'
f_q2 = 'C:/Users/ykqin3/Documents/Loomy Workspace/.loomy-attachments/ses_071ded120ffeOMnlYKgOSIt1Xp/1785086444219-1-_-2026_Q2_.xlsx'
f_annual = 'C:/Users/ykqin3/Documents/Loomy Workspace/.loomy-attachments/ses_071ded120ffeOMnlYKgOSIt1Xp/1785086444221-2-_-_2026_.xlsx'

def extract_q1_data(wb):
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        sup_score = None
        sup_rating = None
        sup_comment = None
        sup_name = None
        emp_score = None
        emp_rating = None
        tasks = []
        for row in rows:
            if not row:
                continue
            if row[0] == '总评':
                continue
            if row[0] == '评价角色' or row[0] == '考核上级':
                if len(row) > 3 and row[1] and '上级' in str(row[0]):
                    sup_score = row[2] if len(row) > 2 else None
                    sup_rating = row[3] if len(row) > 3 else None
                    sup_comment = row[4] if len(row) > 4 else None
                    sup_name = row[1] if len(row) > 1 else None
                if len(row) > 3 and row[1] and ('员工' in str(row[0]) or '自评' in str(row[0])):
                    emp_score = row[2] if len(row) > 2 else None
                    emp_rating = row[3] if len(row) > 3 else None
            elif row[0] and str(row[0]).strip() not in ['指标名称', '', 'None', '评价角色', '考核上级', '员工']:
                task_name = str(row[0]).strip()
                weight = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                emp = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                other = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                sup = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                completion = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                if task_name:
                    tasks.append({'name': task_name, 'weight': weight, 'emp_score': emp, 'other_score': other, 'sup_score': sup, 'completion': completion})
        data[sheet_name] = {'tasks': tasks, 'emp_score': emp_score, 'emp_rating': emp_rating, 'sup_score': sup_score, 'sup_rating': sup_rating, 'sup_comment': sup_comment, 'sup_name': sup_name}
    return data

def extract_annual_data(wb):
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        dept = None
        position = None
        tasks = []
        for row in rows:
            if not row:
                continue
            if len(row) > 6 and str(row[5]) == '部门':
                dept = str(row[6]) if len(row) > 6 and row[6] else None
            if len(row) > 8 and str(row[7]) == '职位':
                position = str(row[8]) if len(row) > 8 and row[8] else None
            if len(row) > 3 and row[2] and str(row[2]).strip():
                task_name = str(row[2]).strip()
                if task_name in ['指标名称', '序号', '评价标准', '实际完成说明', '自评得分', '上级评分', 'EMT审议评分']:
                    continue
                weight = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                if weight and ('%' in weight or weight.replace('.', '').replace(',', '').isdigit()):
                    tasks.append({'name': task_name, 'weight': weight})
        data[sheet_name] = {'dept': dept, 'position': position, 'tasks': tasks}
    return data

def extract_q2_data(wb):
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        sup_score = None
        sup_rating = None
        sup_comment = None
        sup_name = None
        tasks = []
        for row in rows:
            if not row:
                continue
            if row[0] == '总评':
                continue
            if row[0] == '评价角色' or row[0] == '考核上级':
                if len(row) > 3 and row[1] and '上级' in str(row[0]):
                    sup_score = row[2] if len(row) > 2 else None
                    sup_rating = row[3] if len(row) > 3 else None
                    sup_comment = row[4] if len(row) > 4 else None
                    sup_name = row[1] if len(row) > 1 else None
            elif row[0] and str(row[0]).strip() not in ['指标名称', '', 'None', '评价角色', '考核上级', '员工']:
                task_name = str(row[0]).strip()
                weight = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                emp = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                other = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                sup = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                completion = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                if task_name:
                    tasks.append({'name': task_name, 'weight': weight, 'emp_score': emp, 'other_score': other, 'sup_score': sup, 'completion': completion})
        data[sheet_name] = {'tasks': tasks, 'sup_score': sup_score, 'sup_rating': sup_rating, 'sup_comment': sup_comment, 'sup_name': sup_name}
    return data

def parse_q1_focus(comment):
    if not comment:
        return ''
    comment = str(comment)
    match = re.search(r'\(2\)下周期工作重点关注[：:]\s*(.+?)(?=\(3\)|\n\(3\)|未来个人发展|$)', comment, re.DOTALL)
    if match:
        return match.group(1).strip()
    match = re.search(r'下周期工作重点关注[：:]\s*(.+?)(?=\(3\)|未来个人发展|$)', comment, re.DOTALL)
    if match:
        return match.group(1).strip()
    return ''

def extract_key_sentence(completion):
    """从完成说明中提取关键句（包含结果、进展、数据等）"""
    if not completion:
        return ''
    
    text = str(completion)
    # 按句号、分号、换行分割成句子
    sentences = re.split(r'[。；;\n]', text)
    
    # 优先选择包含数字、百分比、完成、达成、推进等关键词的句子
    key_keywords = ['完成', '达成', '推进', '实现', '取得', '提升', '降低', '增长', '通过', '上线', '发布', '立项']
    
    for sentence in sentences:
        sentence = sentence.strip()
        if not sentence:
            continue
        # 跳过以序号开头的句子（如"1、"）
        if re.match(r'^\d+[,\.、]', sentence):
            continue
        # 检查是否包含数字（如百分比、数量等）
        has_numbers = bool(re.search(r'\d+', sentence))
        # 检查是否包含关键词
        has_keywords = any(kw in sentence for kw in key_keywords)
        
        if has_numbers or has_keywords:
            # 截取前50字
            return sentence[:50]
    
    # 如果没有找到关键句，返回第一句（前50字）
    first_sentence = sentences[0].strip() if sentences else ''
    return first_sentence[:50]

def match_tasks(task_name, quarterly_tasks):
    """更智能的任务匹配：不仅比较名称，还比较内容关键词"""
    # 首先尝试名称匹配
    for task in quarterly_tasks:
        if task['name'] in task_name or task_name in task['name'] or (len(task_name) > 5 and task_name[:5] in task['name']) or (len(task['name']) > 5 and task['name'][:5] in task_name):
            return task
    
    # 如果名称不匹配，尝试关键词匹配
    annual_keywords = set(task_name.split())
    for task in quarterly_tasks:
        task_text = task['name'] + ' ' + str(task.get('completion', ''))
        if any(kw in task_text for kw in annual_keywords if len(kw) > 1):
            return task
    
    return None

def get_rating_class(rating):
    mapping = {'A': 'rating-A', 'B': 'rating-B', 'C': 'rating-C', 'D': 'rating-D'}
    return mapping.get(rating, 'rating-C')

def get_text_class(rating):
    mapping = {'A': 'text-A', 'B': 'text-B', 'C': 'text-C', 'D': 'text-D'}
    return mapping.get(rating, 'text-C')

def get_avatar_class(rating):
    mapping = {'A': 'avatar-A', 'B': 'avatar-B', 'C': 'avatar-C', 'D': 'avatar-D'}
    return mapping.get(rating, 'avatar-C')

def generate_highlight(name, q2_info, q1_info):
    """综合高分项指标、员工自评和上级评语提炼关键亮点"""
    highlight_parts = []
    
    # 1. 找到得分最高的任务（高分项指标）
    q2_tasks = q2_info.get('tasks', [])
    best_task = None
    best_score = 0
    for task in q2_tasks:
        if task['name'] == '员工':
            continue
        try:
            s = float(task.get('sup_score', 0))
            if s > best_score:
                best_score = s
                best_task = task
        except:
            pass
    
    if best_task and best_score >= 110:
        highlight_parts.append(f"{best_task['name']}({best_task['sup_score']}分)")
    
    # 2. 提取上级评语中的"做得好的"部分
    sup_comment = q2_info.get('sup_comment', '')
    if sup_comment:
        # 尝试提取"做得好的"后面的内容
        match = re.search(r'做得好的[：:](.+?)(?=\(|$|\n\(|。|;)', str(sup_comment), re.DOTALL)
        if match:
            good_part = match.group(1).strip().replace('\n', '').replace('  ', ' ')[:60]
            if good_part and len(good_part) > 5:
                highlight_parts.append(good_part)
    
    # 3. 提取Q2完成说明中的关键信息（员工业绩亮点）
    if best_task:
        completion = best_task.get('completion', '')
        if completion:
            # 提取完成说明的第一句话
            first_sentence = str(completion).split('。')[0].strip().replace('\n', '').replace('  ', ' ')[:50]
            if first_sentence and len(first_sentence) > 10:
                pass  # 不单独添加，避免太长
    
    # 组合提炼
    if highlight_parts:
        result = '；'.join(highlight_parts[:2])  # 最多2个要点
        return result[:100]  # 限制长度
    
    # 备选：提取Q2最高得分任务的名称
    if best_task:
        return f"{best_task['name']}表现突出，得分{best_task['sup_score']}"
    
    return ''

# Read data
wb_q1 = openpyxl.load_workbook(f_q1)
q1_data = extract_q1_data(wb_q1)

wb_q2 = openpyxl.load_workbook(f_q2)
q2_data = extract_q2_data(wb_q2)

wb_annual = openpyxl.load_workbook(f_annual)
annual_data = extract_annual_data(wb_annual)

# Fix departments and supervisor names
if '胡素琴' in annual_data:
    annual_data['胡素琴']['dept'] = '组人部'
if '蔡尚' in annual_data:
    annual_data['蔡尚']['dept'] = '财务中心'
# Fix supervisor name for Dong Bin
for name in q2_data:
    if q2_data[name].get('sup_name', '').startswith('吴骏华'):
        q2_data[name]['sup_name'] = '吴骏华'

all_names = sorted(q2_data.keys())

# Calculate stats
scores = []
for name in all_names:
    info = q2_data[name]
    if info['sup_score']:
        try:
            s = float(str(info['sup_score']).replace(',', ''))
            scores.append((name, s, info['sup_rating']))
        except:
            pass

scores.sort(key=lambda x: x[1], reverse=True)

rating_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
for name, score, rating in scores:
    if rating in rating_counts:
        rating_counts[rating] += 1

avg_score = round(sum(s[1] for s in scores) / len(scores), 2) if scores else 0
max_name, max_score, _ = scores[0] if scores else ('', 0, '')
min_name, min_score, _ = scores[-1] if scores else ('', 0, '')

# Generate person cards (no score bars)
person_cards = []
for name, score, rating in scores:
    info = q2_data[name]
    dept = annual_data.get(name, {}).get('dept') or ''
    tasks_html = []
    for task in info['tasks']:
        if task['name'] == '员工':
            continue
        task_name = task['name']
        weight = task['weight']
        emp = task['emp_score']
        sup = task['sup_score']
        # Determine text color based on score
        text_class = get_text_class(rating)
        tasks_html.append(f"""<tr><td class="task-name" title="{escape(task_name)}">{escape(task_name)}</td><td class="task-weight">{escape(weight)}</td><td>{escape(emp)}</td><td><span class="task-score {text_class}">{escape(sup)}</span></td></tr>""")
    card = f"""<div class="person-card"><div class="person-head"><div class="person-avatar {get_avatar_class(rating)}">{name[0]}</div><div class="person-meta"><div class="name">{name}</div><div class="role">{escape(dept)}</div></div><div class="person-score"><span class="score-label">上级评级</span><span class="rating {get_rating_class(rating)}">{rating}</span><div class="score {get_text_class(rating)}">{score}</div></div></div><table class="task-table"><tr><th>任务</th><th>权重</th><th>自评</th><th>上级</th></tr>{''.join(tasks_html)}</table></div>"""
    person_cards.append(card)

# Generate ranking table rows
ranking_rows = []
for i, (name, score, rating) in enumerate(scores):
    info = q2_data[name]
    dept = annual_data.get(name, {}).get('dept') or ''
    sup_name = info.get('sup_name') or ''
    
    # Generate highlight from high-scoring tasks, self-eval, and supervisor comment
    highlight = generate_highlight(name, info, q1_data.get(name, {}))
    
    rank_color = '#165dff' if i < 5 else '#f7ba1e'
    bg_style = 'background:#f7f8fa;' if i % 2 == 0 else ''
    ranking_rows.append(f"""<tr style="border-top:1px solid #f2f3f5;{bg_style}"><td style="padding:7px 12px;font-weight:700;color:{rank_color};white-space:nowrap;width:36px;">{i+1}</td><td style="padding:7px 12px;font-weight:600;white-space:nowrap;width:48px;">{name}</td><td style="padding:7px 12px;color:#4e5969;white-space:nowrap;width:80px;font-size:12px;">{escape(dept)}</td><td style="padding:7px 12px;color:#4e5969;white-space:nowrap;width:72px;font-size:12px;">{escape(sup_name)}</td><td style="padding:7px 12px;font-weight:700;white-space:nowrap;width:48px;">{score}</td><td style="padding:7px 12px;white-space:nowrap;width:48px;"><span class="rating {get_rating_class(rating)}" style="padding:1px 10px;border-radius:8px;color:#fff;font-size:12px;">{rating}</span></td><td style="padding:7px 12px;color:#4e5969;font-size:13px;">{escape(highlight)}</td></tr>""")

# Generate annual progress cards with Q1/Q2 completion status
annual_cards = []
for name in all_names:
    score = 0
    rating = 'C'
    for n, s, r in scores:
        if n == name:
            score = s
            rating = r
            break
    dept = annual_data.get(name, {}).get('dept') or ''
    q1_rating = q1_data.get(name, {}).get('sup_rating', 'N/A') or 'N/A'
    annual_tasks = annual_data.get(name, {}).get('tasks', [])
    q2_tasks = q2_data.get(name, {}).get('tasks', [])
    q1_tasks = q1_data.get(name, {}).get('tasks', [])
    
    if score >= 115:
        risk = '低'
        risk_color = '#00b42a'
        risk_bg = '#e8f8e8'
        risk_label = ' 进度超前'
    elif score >= 110:
        risk = '低'
        risk_color = '#165dff'
        risk_bg = '#e8f3ff'
        risk_label = '🔵 正常推进'
    elif score >= 105:
        risk = '中低'
        risk_color = '#f7ba1e'
        risk_bg = '#fff7e8'
        risk_label = '🟡 需关注'
    else:
        risk = '中'
        risk_color = '#f53f3f'
        risk_bg = '#ffece8'
        risk_label = '🔴 风险预警'
    
    # Annual tasks with Q1/Q2 completion status
    annual_tasks_html = []
    for t in annual_tasks:
        task_name = t['name']
        # Find Q1 and Q2 completion status using smart matching
        q1_task = match_tasks(task_name, q1_tasks)
        q2_task = match_tasks(task_name, q2_tasks)
        
        q1_status = ''
        q2_status = ''
        
        if q1_task:
            completion = q1_task.get('completion', '')
            if completion:
                key_line = extract_key_sentence(completion)
                if key_line:
                    q1_status = f"Q1：{key_line}"
        
        if q2_task:
            completion = q2_task.get('completion', '')
            if completion:
                key_line = extract_key_sentence(completion)
                if key_line:
                    q2_status = f"Q2：{key_line}"
        
        status_text = ''
        if q1_status or q2_status:
            status_text = ' '.join(filter(None, [q1_status, q2_status]))
            annual_tasks_html.append(f"<li style='margin-bottom:4px;'><strong>{escape(task_name)}</strong><br/><span style='color:#86909c;font-size:12px;'>{escape(status_text)}</span></li>")
        else:
            annual_tasks_html.append(f"<li style='margin-bottom:4px;'><strong>{escape(task_name)}</strong></li>")
    
    # Tasks not reflected in quarterly - check both name and completion content
    q2_task_data = {task['name']: task.get('completion', '') for task in q2_tasks if task.get('name')}
    q1_task_data = {task['name']: task.get('completion', '') for task in q1_tasks if task.get('name')}
    all_quarterly_tasks = set(q2_task_data.keys()) | set(q1_task_data.keys())
    
    not_in_quarterly = []
    for t in annual_tasks:
        task_name = t['name']
        found = False
        # First check by name matching
        for q_task_name in all_quarterly_tasks:
            if task_name in q_task_name or q_task_name in task_name or (len(task_name) > 5 and task_name[:5] in q_task_name) or (len(q_task_name) > 5 and q_task_name[:5] in task_name):
                found = True
                break
        # Also check if annual task keywords appear in quarterly task completions
        if not found:
            annual_keywords = task_name.split()[:3]  # First 3 keywords
            for q_task_name, q_completion in {**q2_task_data, **q1_task_data}.items():
                q_text = (q_task_name + ' ' + str(q_completion))[:200]
                if any(kw in q_text for kw in annual_keywords if len(kw) > 1):
                    found = True
                    break
        if not found:
            # Skip "组织绩效承接" type tasks
            if '组织绩效' not in task_name and '承接' not in task_name:
                not_in_quarterly.append(task_name)
    
    not_in_quarterly_html = ''
    if not_in_quarterly:
        not_in_quarterly_html = f"""<div style="margin-top:8px;font-size:12px;color:#4e5969;line-height:1.8;"><strong>未在季度考核体现的年度任务：</strong><ul style="margin:0;padding-left:20px;">{''.join([f'<li>{escape(t)}</li>' for t in not_in_quarterly[:2]])}</ul></div>"""
    
    # Stalled tasks - use sup_score only, threshold < 100
    stalled_tasks = []
    for task in q2_tasks:
        try:
            s = float(task['sup_score']) if task['sup_score'] else 100
            if s < 100:
                stalled_tasks.append(f"{task['name']} (得分{s})")
        except:
            pass
    
    stalled_html = ''
    if stalled_tasks:
        stalled_html = f"""<div style="margin-top:8px;font-size:12px;color:#4e5969;line-height:1.8;"><strong>推进不顺利的任务：</strong><ul style="margin:0;padding-left:20px;">{''.join([f'<li>{escape(t)}</li>' for t in stalled_tasks[:2]])}</ul></div>"""
    
    card = f"""<div style="background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.06);border-left:4px solid {risk_color};"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;"><span style="font-size:16px;font-weight:600;">{name}</span><span style="font-size:12px;background:{risk_bg};color:{risk_color};padding:2px 10px;border-radius:6px;font-weight:700;">{risk_label}</span></div><div style="font-size:12px;color:#86909c;margin-bottom:12px;">{escape(dept)} · Q1:{q1_rating} · Q2:{rating}</div><div style="font-size:13px;color:#4e5969;line-height:1.8;"><div style="font-weight:600;margin-bottom:8px;color:#1d2129;">📋 年度目标完成情况：</div><ul style="margin:0;padding-left:20px;margin-bottom:12px;">{''.join(annual_tasks_html) if annual_tasks_html else '<li>详见年度考核表</li>'}</ul>{not_in_quarterly_html}{stalled_html}<div style="margin-top:12px;padding:8px 12px;background:{risk_bg};border-radius:6px;font-size:12px;"><strong>风险等级：{risk}</strong> — {'年度任务正常推进' if rating in ['A', 'B'] else '需关注下半年目标达成'}</div></div></div>"""
    annual_cards.append(card)

# Generate Q1→Q2 closure cards (unchanged)
closure_cards = []
for name in all_names:
    q1_info = q1_data.get(name, {})
    q2_info = q2_data.get(name, {})
    dept = annual_data.get(name, {}).get('dept') or ''
    q2_rating = q2_info.get('sup_rating', 'N/A') or 'N/A'
    
    comment = q1_info.get('sup_comment', '')
    focus = parse_q1_focus(comment)
    
    special_note = ''
    if name == '沈明花':
        special_note = '（Q1产假）'
        focus = '产假期间，不参与Q1考核'
    elif name == '沈明星':
        special_note = '（Q1试用期）'
        focus = '试用期期间，不参与Q1考核'
    
    q1_score_val = q1_info.get('sup_score', 'N/A') or 'N/A'
    q2_score_val = q2_info.get('sup_score', 'N/A') or 'N/A'
    
    if focus and focus.strip() and not special_note:
        try:
            q1_s = float(str(q1_score_val).replace(',', ''))
            q2_s = float(str(q2_score_val).replace(',', ''))
            if q2_s > q1_s:
                status = '✅ 已闭环'
                status_color = '#00b42a'
            elif q2_s == q1_s:
                status = '️ 持平'
                status_color = '#f7ba1e'
            else:
                status = ' 未闭环'
                status_color = '#f53f3f'
        except:
            status = ' 待评估'
            status_color = '#86909c'
    else:
        if special_note:
            status = '️ 特殊说明'
            status_color = '#86909c'
        else:
            status = '️ 无Q1数据'
            status_color = '#86909c'
    
    if not focus or not focus.strip():
        focus = '（Q1无明确重点关注）'
    
    # Q2 completion summary
    q2_tasks = q2_info.get('tasks', [])
    q2_completion_html = ''
    if q2_tasks:
        task_items = []
        for task in q2_tasks[:3]:
            if task['name'] != '员工':
                task_name = task['name']
                task_score = task.get('sup_score', 'N/A')
                task_items.append(f"{task_name}({task_score})")
        q2_completion_html = '；'.join(task_items)
    
    card = f"""<div style="background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.06);border-left:4px solid {status_color};"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;"><span style="font-size:16px;font-weight:600;">{name}{escape(special_note)} <span style="font-size:11px;background:{status_color}20;color:{status_color};padding:1px 8px;border-radius:6px;margin-left:8px;">Q2:{q2_rating}</span></span><span style="font-size:12px;background:{status_color}20;color:{status_color};padding:2px 10px;border-radius:6px;font-weight:700;">{status}</span></div><div style="font-size:12px;color:#86909c;margin-bottom:8px;">{escape(dept)}</div><div style="font-size:13px;color:#4e5969;line-height:2;"><strong>📌 Q1上级要求：</strong><br>&nbsp;&nbsp;{escape(focus[:150]) if len(focus) > 150 else escape(focus)}<br><br><strong>📊 Q2完成情况：</strong><br>&nbsp;&nbsp;{escape(q2_completion_html) if q2_completion_html else '详见Q2考核指标'}<br><br><strong>🎯 闭环评价：</strong><br>&nbsp;&nbsp;{status}</div></div>"""
    closure_cards.append(card)

# Build full HTML
html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>职能组 · 2026 Q2 绩效重点任务完成进度看板</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Microsoft YaHei", "PingFang SC", sans-serif; background: #f0f2f5; color: #1d2129; padding: 24px 32px; }}
  .header {{ margin-bottom: 28px; }}
  .header h1 {{ font-size: 26px; font-weight: 700; letter-spacing: 1px; }}
  .header .sub {{ display: flex; justify-content: space-between; align-items: center; margin-top: 6px; color: #4e5969; font-size: 14px; }}
  .header .sub .badge {{ display: inline-block; background: #165dff; color: #fff; padding: 2px 14px; border-radius: 12px; font-size: 12px; font-weight: 600; }}
  .legend {{ display: flex; gap: 20px; margin-bottom: 24px; flex-wrap: wrap; }}
  .legend-item {{ display: flex; align-items: center; gap: 6px; font-size: 13px; color: #4e5969; }}
  .legend-dot {{ width: 14px; height: 14px; border-radius: 4px; }}
  .dot-a {{ background: #00b42a; }}
  .dot-b {{ background: #165dff; }}
  .dot-c {{ background: #f7ba1e; }}
  .dot-d {{ background: #f53f3f; }}
  .summary-row {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 28px; }}
  .summary-card {{ background: #fff; border-radius: 12px; padding: 18px 20px; box-shadow: 0 1px 4px rgba(0,0,0,.06); }}
  .summary-card .label {{ font-size: 13px; color: #86909c; }}
  .summary-card .value {{ font-size: 28px; font-weight: 700; margin-top: 4px; }}
  .summary-card .value .unit {{ font-size: 14px; font-weight: 400; color: #86909c; margin-left: 4px; }}
  .dept-section {{ margin-bottom: 32px; }}
  .dept-header {{ display: flex; align-items: center; gap: 12px; margin-bottom: 14px; }}
  .dept-header h2 {{ font-size: 18px; font-weight: 600; }}
  .dept-header .sup {{ font-size: 13px; color: #86909c; background: #f2f3f5; padding: 2px 12px; border-radius: 10px; }}
  .dept-header .dept-count {{ font-size: 13px; color: #86909c; margin-left: auto; }}
  .person-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(420px, 1fr)); gap: 16px; }}
  .person-card {{ background: #fff; border-radius: 12px; box-shadow: 0 1px 4px rgba(0,0,0,.06); overflow: hidden; transition: box-shadow .2s; }}
  .person-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,.1); }}
  .person-head {{ display: flex; align-items: center; gap: 14px; padding: 16px 20px 12px; border-bottom: 1px solid #f2f3f5; }}
  .person-avatar {{ width: 40px; height: 40px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 18px; font-weight: 700; color: #fff; flex-shrink: 0; }}
  .person-meta {{ flex: 1; }}
  .person-meta .name {{ font-size: 16px; font-weight: 600; }}
  .person-meta .role {{ font-size: 12px; color: #86909c; margin-top: 2px; }}
  .person-score {{ display: flex; align-items: center; gap: 8px; }}
  .person-score .score {{ font-size: 22px; font-weight: 700; }}
  .person-score .rating {{ display: inline-block; font-size: 13px; font-weight: 600; padding: 1px 10px; border-radius: 8px; color: #fff; }}
  .person-score .score-label {{ font-size: 11px; color: #86909c; display: block; margin-top: 1px; width: 100%; text-align: center; }}
  .task-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .task-table th {{ text-align: left; padding: 10px 20px 6px; font-weight: 500; color: #86909c; font-size: 11px; letter-spacing: .5px; }}
  .task-table td {{ padding: 7px 20px; border-bottom: 1px solid #f7f8fa; vertical-align: top; }}
  .task-table td:last-child {{ text-align: right; }}
  .task-table tr:last-child td {{ border-bottom: none; }}
  .task-name {{ max-width: 250px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  .task-weight {{ color: #86909c; font-size: 12px; }}
  .task-score {{ font-weight: 600; min-width: 36px; display: inline-block; }}
  .rating-A {{ background: #00b42a; }}
  .rating-B {{ background: #165dff; }}
  .rating-C {{ background: #f7ba1e; }}
  .rating-D {{ background: #f53f3f; }}
  .text-A {{ color: #00b42a; }}
  .text-B {{ color: #165dff; }}
  .text-C {{ color: #f7ba1e; }}
  .text-D {{ color: #f53f3f; }}
  .avatar-A {{ background: #00b42a; }}
  .avatar-B {{ background: #165dff; }}
  .avatar-C {{ background: #f7ba1e; }}
  .avatar-D {{ background: #f53f3f; }}
  .footer {{ text-align: center; color: #c9cdd4; font-size: 12px; padding: 32px 0 8px; border-top: 1px solid #e5e6eb; margin-top: 16px; }}
</style>
</head>
<body>

<div class="header">
  <h1>职能组 · 2026 Q2 绩效重点任务完成进度看板</h1>
  <div class="sub">
    <span>用于 Q2 绩效分档评级参考 · 数据来源：中层干部年度 &amp; Q1/Q2 考核表</span>
    <span class="badge">2026年7月 · 半年度</span>
  </div>
</div>

<div class="legend">
  <div class="legend-item"><div class="legend-dot dot-a"></div>A 档 (130~150) 卓越</div>
  <div class="legend-item"><div class="legend-dot dot-b"></div>B 档 (110~130) 超出预期</div>
  <div class="legend-item"><div class="legend-dot dot-c"></div>C 档 (90~110) 正常完成</div>
  <div class="legend-item"><div class="legend-dot dot-d"></div>D 档 (60~90) 需改进</div>
</div>

<div class="summary-row">
  <div class="summary-card"><div class="label">参评干部</div><div class="value">{len(scores)} <span class="unit">人</span></div></div>
  <div class="summary-card"><div class="label">Q2 上级评分均值</div><div class="value">{avg_score} <span class="unit">分</span></div></div>
  <div class="summary-card"><div class="label">A 档人数</div><div class="value">{rating_counts['A']} <span class="unit">人</span></div></div>
  <div class="summary-card"><div class="label">B 档人数</div><div class="value">{rating_counts['B']} <span class="unit">人</span></div></div>
  <div class="summary-card"><div class="label">C 档人数</div><div class="value">{rating_counts['C']} <span class="unit">人</span></div></div>
  <div class="summary-card"><div class="label">最高分</div><div class="value">{max_score} <span class="unit">分 · {max_name}</span></div></div>
  <div class="summary-card"><div class="label">最低分</div><div class="value">{min_score} <span class="unit">分 · {min_name}</span></div></div>
</div>

<div class="dept-section">
  <div class="dept-header">
    <h2>👥 全员Q2考核指标</h2>
    <span class="dept-count">按上级评分排序 · 共 {len(scores)} 人</span>
  </div>
  <div class="person-grid">
    {''.join(person_cards)}
  </div>
</div>

<div class="dept-section">
  <div class="dept-header">
    <h2>📊 Q2 全员总分排名</h2>
    <span class="dept-count">按上级评分排序</span>
  </div>
  <div style="background:#fff;border-radius:12px;box-shadow:0 1px 4px rgba(0,0,0,.06);overflow:hidden;">
    <table style="width:100%;border-collapse:collapse;font-size:13px;">
      <tr style="background:#f7f8fa;">
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:36px;">排名</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:48px;">姓名</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:80px;">所在部门</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:72px;">考核上级</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:48px;">Q2评分</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:48px;">评级</th>
        <th style="padding:8px 12px;text-align:left;font-weight:500;color:#86909c;width:50%;">关键亮点</th>
      </tr>
      {''.join(ranking_rows)}
    </table>
  </div>
</div>

<div class="dept-section">
  <div class="dept-header">
    <h2>🎯 年度重点任务完成进度分析（H1）</h2>
    <span class="dept-count">结合年度目标 vs Q1+Q2 实际完成</span>
  </div>
  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:16px;">
    {''.join(annual_cards)}
  </div>
</div>

<div class="dept-section">
  <div class="dept-header">
    <h2>🔄 Q1上级重点关注 → Q2完成闭环</h2>
    <span class="dept-count">上级在Q1提出的下周期重点工作，Q2是否落实</span>
  </div>
  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:16px;">
    {''.join(closure_cards)}
  </div>
</div>

<div class="footer">职能组 · 2026 Q2 绩效看板 · 数据截止 2026年7月</div>

</body>
</html>
"""

# Save to local file
output_path = 'C:/Users/ykqin3/Documents/Loomy Workspace/func-group-dashboard-final.html'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Final HTML saved to: {output_path}")
print(f"File size: {len(html)} bytes")
print(f"Total cadres: {len(scores)}")
